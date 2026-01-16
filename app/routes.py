from flask import Blueprint, render_template, request, jsonify, send_file, session, redirect, url_for
import csv, io, datetime, openpyxl
from openpyxl.styles import Font, Alignment, NamedStyle
import pandas as pd
from .utils import format_taxpayer_id
import sqlite3
import os

# path for the sqlite database file (stored in project root)
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'clients.db')

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tin TEXT UNIQUE,
            corporate_name TEXT,
            last_name TEXT,
            first_name TEXT,
            middle_name TEXT,
            trade_name TEXT,
            address1 TEXT,
            address2 TEXT,
            rdo TEXT,
            last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

# initialize DB when module is loaded
init_db()

main = Blueprint('main', __name__)

VALID_USERNAME = "admin"
VALID_PASSWORD = "xdt12345"

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('main.login'))
        return f(*args, **kwargs)
    return decorated


@main.route('/', methods=['GET', 'POST'])
def login():
    # Clear any existing session data on login
    session.pop('latest_sums_purchases', None)
    session.pop('latest_sums_sales', None)
    session.pop('latest_sums_quarterly', None)
    session.pop('latest_sums_annual', None)
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()

        if username == VALID_USERNAME and password == VALID_PASSWORD:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('main.dashboard'))
        else:
            return render_template('login.html', error="Invalid credentials")

    return render_template('login.html')

@main.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('main.login'))
    username = session.get('username', 'User')
    return render_template('dashboard.html', username=username)

@main.route('/vat-relief')
def vat_relief():
    if 'username' not in session:
        return redirect(url_for('main.login'))
    return render_template('vat_relief.html')

@main.route('/sales')
def sales():
    if 'username' not in session:
        return redirect(url_for('main.login'))
    return render_template('sales.html')

@main.route('/sawt')
def sawt():
    if 'username' not in session:
        return redirect(url_for('main.login'))
    return render_template('sawt.html')

@main.route('/qap-quarterly')
def qap_quarterly():
    if 'username' not in session:
        return redirect(url_for('main.login'))
    return render_template('quarterly.html')

@main.route('/qap-annual')
def qap_annual():
    if 'username' not in session:
        return redirect(url_for('main.login'))
    return render_template('annual.html')


@main.route('/logout')
def logout():
    session.clear()  # Clear all session data
    return redirect(url_for('main.login'))

@main.route('/')
@login_required
def index():
    return render_template('index.html')


@main.route('/upload_csv', methods=['POST'])
def upload_csv():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        content = f.stream.read().decode('ISO-8859-1').splitlines()
        reader = csv.reader(content)
        header = next(reader, None)

        consolidated = {}
        sums = {'exempt': 0, 'zero': 0, 'serv': 0, 'cap': 0, 'other': 0, 'inputtax': 0}

        for row in reader:
            if not row:
                continue
            key = row[0].replace('-', '').upper().ljust(9)[:9]
            if key not in consolidated:
                consolidated[key] = ['D', 'P', key] + [c.upper() for c in row[1:7]] + [0.0]*6

            for i in range(7, 13):
                try:
                    consolidated[key][i+2] += float(row[i])
                except:
                    pass

        sorted_data = sorted(consolidated.values(), key=lambda x: (x[3], x[4]))
        rows_out = []

        for vals in sorted_data:
            formatted = []
            for i, v in enumerate(vals):
                if i < 2:
                    formatted.append(str(v))
                elif i < 3:
                    formatted.append(f'"{v}"')
                elif 3 <= i <= 8:
                    formatted.append(f'"{v}"' if str(v).strip() else v)
                elif 9 <= i <= 14:
                    formatted.append(f"{float(v):.2f}")
                else:
                    formatted.append(str(v))

            try:
                sums['exempt'] += float(vals[9])
                sums['zero'] += float(vals[10])
                sums['serv'] += float(vals[11])
                sums['cap'] += float(vals[12])
                sums['other'] += float(vals[13])
                sums['inputtax'] += float(vals[14])
            except:
                pass

            rows_out.append(','.join(formatted))

        # Store sums in session instead of global variable
        session['latest_sums_purchases'] = sums
        return jsonify({'rows': rows_out, 'sums': sums})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main.route('/upload_sales', methods=['POST'])
def upload_sales():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        content = f.stream.read().decode('ISO-8859-1').splitlines()
        reader = csv.reader(content)
        header = next(reader, None)

        rows_out = []
        sums = {'exempt': 0, 'zero': 0, 'taxable': 0, 'outputvat': 0}

        for row in reader:
            if not row:
                continue

            formatted = ['D', 'S']
            for i, v in enumerate(row):
                v = v.strip()
                if i <= 7:
                    formatted.append(f'"{v}"' if v else '')
                else:
                    try:
                        formatted.append(f"{float(v):.2f}")
                    except:
                        formatted.append(v)

            try:
                sums['exempt'] += float(row[7]) if row[7] else 0
                sums['zero'] += float(row[8]) if row[8] else 0
                sums['taxable'] += float(row[9]) if row[9] else 0
                sums['outputvat'] += float(row[10]) if row[10] else 0
            except:
                pass

            rows_out.append(','.join(formatted))

        # Store sums in session instead of global variable
        session['latest_sums_sales'] = sums
        return jsonify({'rows': rows_out, 'sums': sums})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main.route('/upload_quarterly', methods=['POST'])
def upload_quarterly():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        content = f.stream.read().decode('ISO-8859-1').splitlines()
        reader = csv.reader(content)
        header = next(reader, None)

        rows_out = []
        sums = {'total': 0}

        for row in reader:
            if not row:
                continue

            # Process quarterly data
            # You'll need to adjust this based on your actual quarterly CSV format
            formatted = ['D', 'Q']  # Q for quarterly
            
            # Example processing - adjust based on your needs
            for i, v in enumerate(row):
                v = v.strip()
                if i <= 5:  # Adjust this based on your CSV format
                    formatted.append(f'"{v}"' if v else '')
                else:
                    try:
                        formatted.append(f"{float(v):.2f}")
                        if i == 6:  # Assuming column 7 is the withholding amount
                            sums['total'] += float(v)
                    except:
                        formatted.append(v)

            rows_out.append(','.join(formatted))

        session['latest_sums_quarterly'] = sums
        return jsonify({'rows': rows_out, 'sums': sums})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@main.route('/upload_annual', methods=['POST'])
def upload_annual():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400

    try:
        content = f.stream.read().decode('ISO-8859-1').splitlines()
        reader = csv.reader(content)
        header = next(reader, None)

        rows_out = []
        sums = {'total': 0}

        for row in reader:
            if not row:
                continue

            # Process annual data
            # You'll need to adjust this based on your actual annual CSV format
            formatted = ['D', 'A']  # A for annual
            
            # Example processing - adjust based on your needs
            for i, v in enumerate(row):
                v = v.strip()
                if i <= 5:  # Adjust this based on your CSV format
                    formatted.append(f'"{v}"' if v else '')
                else:
                    try:
                        formatted.append(f"{float(v):.2f}")
                        if i == 6:  # Assuming column 7 is the withholding amount
                            sums['total'] += float(v)
                    except:
                        formatted.append(v)

            rows_out.append(','.join(formatted))

        session['latest_sums_annual'] = sums
        return jsonify({'rows': rows_out, 'sums': sums})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    
@main.route('/submit_sales', methods=['POST'])
def submit_sales():
    data = request.get_json() or {}

    def q(v): return f'"{(v or "").upper()}"'

    client_tin = (data.get('client_tin') or '').replace('-', '')
    rdo = (data.get('rdo') or '').zfill(3)
    period = (data.get('period') or '').upper()

    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO clients (tin, corporate_name, last_name, first_name, middle_name, trade_name, address1, address2, rdo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(tin) DO UPDATE SET
                corporate_name=excluded.corporate_name,
                last_name=excluded.last_name,
                first_name=excluded.first_name,
                middle_name=excluded.middle_name,
                trade_name=excluded.trade_name,
                address1=excluded.address1,
                address2=excluded.address2,
                rdo=excluded.rdo,
                last_updated = CURRENT_TIMESTAMP
        ''', (
            client_tin,
            (data.get('corporate_name') or '').upper(),
            (data.get('last_name') or '').upper(),
            (data.get('first_name') or '').upper(),
            (data.get('middle_name') or '').upper(),
            (data.get('trade_name') or '').upper(),
            (data.get('address1') or '').upper(),
            (data.get('address2') or '').upper(),
            rdo,
        ))
        conn.commit()
    except Exception as e:
        print("DB save error:", e)
    finally:
        try:
            conn.close()
        except:
            pass

    # Get sums from session with default values
    sums = session.get('latest_sums_sales', {
        'exempt': 0, 'zero': 0, 'taxable': 0, 'outputvat': 0
    })

    new_data = [
        'H,S',
        f'"{client_tin}"',
        q(data.get('corporate_name')),
        q(data.get('last_name')),
        q(data.get('first_name')),
        q(data.get('middle_name')),
        q(data.get('trade_name')),
        q(data.get('address1')),
        q(data.get('address2')),
        f"{sums['exempt']:.2f}",
        f"{sums['zero']:.2f}",
        f"{sums['taxable']:.2f}",
        f"{sums['outputvat']:.2f}",
        rdo,
        period,
        "12"
    ]

    return jsonify({'row': ','.join(new_data)})

@main.route('/submit_entry', methods=['POST'])
def submit_entry():
    data = request.get_json() or {}

    def q(v): return f'"{(v or "").upper()}"'

    client_tin = (data.get('client_tin') or '').replace('-', '')
    rdo = (data.get('rdo') or '').zfill(3)
    period = (data.get('period') or '').upper()

    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO clients (tin, corporate_name, last_name, first_name, middle_name, trade_name, address1, address2, rdo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(tin) DO UPDATE SET
                corporate_name=excluded.corporate_name,
                last_name=excluded.last_name,
                first_name=excluded.first_name,
                middle_name=excluded.middle_name,
                trade_name=excluded.trade_name,
                address1=excluded.address1,
                address2=excluded.address2,
                rdo=excluded.rdo,
                last_updated = CURRENT_TIMESTAMP
        ''', (
            client_tin,
            (data.get('corporate_name') or '').upper(),
            (data.get('last_name') or '').upper(),
            (data.get('first_name') or '').upper(),
            (data.get('middle_name') or '').upper(),
            (data.get('trade_name') or '').upper(),
            (data.get('address1') or '').upper(),
            (data.get('address2') or '').upper(),
            rdo,
        ))
        conn.commit()
    except Exception as e:
        print("DB save error:", e)
    finally:
        try:
            conn.close()
        except:
            pass

    # Get sums from session with default values
    sums = session.get('latest_sums_purchases', {
        'exempt': 0, 'zero': 0, 'serv': 0, 'cap': 0, 'other': 0, 'inputtax': 0
    })

    new_data = [
        'H,P',
        f'"{client_tin}"',
        q(data.get('corporate_name')),
        q(data.get('last_name')),
        q(data.get('first_name')),
        q(data.get('middle_name')),
        q(data.get('trade_name')),
        q(data.get('address1')),
        q(data.get('address2')),
        f"{sums['exempt']:.2f}",
        f"{sums['zero']:.2f}",
        f"{sums['serv']:.2f}",
        f"{sums['cap']:.2f}",
        f"{sums['other']:.2f}",
        f"{sums['inputtax']:.2f}",
        f"{sums['inputtax']:.2f}",
        "0.00",
        rdo,
        period,
        "12"
    ]

    return jsonify({'row': ','.join(new_data)})

@main.route('/submit_quarterly', methods=['POST'])
def submit_quarterly():
    data = request.get_json() or {}

    client_tin = (data.get('client_tin') or '').replace('-', '')
    client_branch = (data.get('client_branch') or '').zfill(4)
    period = (data.get('period') or '')

    # Get sums from session with default values
    sums = session.get('latest_sums_quarterly', {'total': 0})

    # Format for quarterly (adjust based on your needs)
    header = f"H,1601EQ,{client_tin},{client_branch},{period}"
    
    # Create data row (adjust based on your needs)
    data_row = f"D,1601EQ,{client_tin},{client_branch},{period},001,{sums['total']:.2f}"

    return jsonify({'header': header, 'data_row': data_row, 'sums': sums})

@main.route('/submit_annual', methods=['POST'])
def submit_annual():
    data = request.get_json() or {}

    client_tin = (data.get('client_tin') or '').replace('-', '')
    client_branch = (data.get('client_branch') or '').zfill(4)
    period = (data.get('period') or '')

    # Get sums from session with default values
    sums = session.get('latest_sums_annual', {'total': 0})

    # Format for annual (adjust based on your needs)
    header = f"H,1604E,{client_tin},{client_branch},{period}"
    
    # Create data row (adjust based on your needs)
    data_row = f"D,1604E,{client_tin},{client_branch},{period},001,{sums['total']:.2f}"

    return jsonify({'header': header, 'data_row': data_row, 'sums': sums})

@main.route('/get_client/<tin>', methods=['GET'])
def get_client(tin):
    tin_norm = (tin or '').replace('-', '').strip()

    if not tin_norm:
        return jsonify({'found': False}), 400

    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            SELECT tin, corporate_name, last_name, first_name, middle_name, trade_name, address1, address2, rdo
            FROM clients WHERE tin = ?
        ''', (tin_norm,))
        row = c.fetchone()
    except Exception as e:
        print("DB lookup error:", e)
        return jsonify({'found': False}), 500
    finally:
        try:
            conn.close()
        except:
            pass

    if not row:
        return jsonify({'found': False})

    keys = ['tin', 'corporate_name', 'last_name', 'first_name', 'middle_name', 'trade_name', 'address1', 'address2', 'rdo']
    return jsonify({'found': True, **dict(zip(keys, row))})



@main.route('/save_dat', methods=['POST'])
def save_dat():
    payload = request.get_json() or {}
    rows = payload.get('rows') or []
    client_tin = (payload.get('client_tin') or '').replace('-', '')
    period = payload.get('period') or ''
    try:
        parts = period.split('/')
        month_year = parts[0] + parts[2]
    except:
        month_year = datetime.datetime.now().strftime('%m%Y')

    default_name = f"{client_tin}P{month_year}.dat" if client_tin else 'output.dat'

    bio = io.StringIO()
    for line in rows:
        bio.write(line + '\n')
    bio.seek(0)

    response = send_file(
        io.BytesIO(bio.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=default_name
    )
    response.headers['X-Filename'] = default_name
    return response

@main.route('/save_dat_sales', methods=['POST'])
def save_dat_sales():
    payload = request.get_json() or {}
    rows = payload.get('rows') or []
    client_tin = (payload.get('client_tin') or '').replace('-', '')
    period = payload.get('period') or ''
    try:
        parts = period.split('/')
        month_year = parts[0] + parts[2]
    except:
        month_year = datetime.datetime.now().strftime('%m%Y')

    default_name = f"{client_tin}S{month_year}.dat" if client_tin else 'output.dat'

    bio = io.StringIO()
    for line in rows:
        bio.write(line + '\n')
    bio.seek(0)

    response = send_file(
        io.BytesIO(bio.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=default_name
    )
    response.headers['X-Filename'] = default_name
    return response

@main.route('/save_dat_quarterly', methods=['POST'])
def save_dat_quarterly():
    payload = request.get_json() or {}
    rows = payload.get('rows') or []
    client_tin = (payload.get('client_tin') or '').replace('-', '')
    period = payload.get('period') or ''
    
    default_name = f"{client_tin}Q{period.replace('/', '')}.dat" if client_tin else 'quarterly.dat'

    bio = io.StringIO()
    for line in rows:
        bio.write(line + '\n')
    bio.seek(0)

    response = send_file(
        io.BytesIO(bio.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=default_name
    )
    response.headers['X-Filename'] = default_name
    return response

@main.route('/save_dat_annual', methods=['POST'])
def save_dat_annual():
    payload = request.get_json() or {}
    rows = payload.get('rows') or []
    client_tin = (payload.get('client_tin') or '').replace('-', '')
    period = payload.get('period') or ''
    
    default_name = f"{client_tin}A{period.replace('/', '')}.dat" if client_tin else 'annual.dat'

    bio = io.StringIO()
    for line in rows:
        bio.write(line + '\n')
    bio.seek(0)

    response = send_file(
        io.BytesIO(bio.getvalue().encode('utf-8')),
        mimetype='text/plain',
        as_attachment=True,
        download_name=default_name
    )
    response.headers['X-Filename'] = default_name
    return response


@main.route('/convert_xlsx', methods=['POST'])
def convert_xlsx():
    payload = request.get_json() or {}
    rows = payload.get('rows') or []
    if not rows:
        return jsonify({'error': 'No rows provided'}), 400

    processed = []
    # Initialize totals for columns F through N (indices 5 to 13 in processed rows)
    column_totals = [0.0] * 9  # F through N = 9 columns
    
    for idx, line in enumerate(rows):
        if idx == 0:
            continue
        cols = [c.strip().strip('"') for c in line.split(',')]
        if len(cols) < 15:
            continue
        try:
            numeric_values = [(float(cols[i]) if cols[i] else 0.0) for i in range(9, 15)]
        except Exception:
            numeric_values = [0.0] * 6

        # Calculate column values
        gross_purchase = sum(numeric_values[:5])  # Column F (6)
        exempt_purchase = numeric_values[0]        # Column G (7)
        zero_rated_purchase = numeric_values[1]    # Column H (8)
        taxable_purchase = sum(numeric_values[2:5]) # Column I (9)
        services_purchase = numeric_values[2]      # Column J (10)
        capital_goods = numeric_values[3]          # Column K (11)
        other_goods = numeric_values[4]            # Column L (12)
        input_tax = numeric_values[5]              # Column M (13)
        gross_taxable_purchase = sum(numeric_values[2:6]) # Column N (14)
        
        # Add to column totals (F through N)
        column_values = [
            gross_purchase,
            exempt_purchase,
            zero_rated_purchase,
            taxable_purchase,
            services_purchase,
            capital_goods,
            other_goods,
            input_tax,
            gross_taxable_purchase
        ]
        
        for i in range(9):
            column_totals[i] += column_values[i]

        formatted_row = [
            cols[16] if len(cols) > 16 else '',
            format_taxpayer_id(cols[2]) if len(cols) > 2 else '',
            cols[3] if len(cols) > 3 else '',
            f"{cols[4]}, {cols[5]} {cols[6]}",
            f"{cols[7]} {cols[8]}",
            gross_purchase,
            exempt_purchase,
            zero_rated_purchase,
            taxable_purchase,
            services_purchase,
            capital_goods,
            other_goods,
            input_tax,
            gross_taxable_purchase
        ]
        processed.append(formatted_row)

    if not processed:
        return jsonify({'error': 'No valid data to convert'}), 400

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header cells
    header_cells = {
        "A1": "PURCHASE TRANSACTION",
        "A2": "RECONCILIATION OF LISTING FOR ENFORCEMENT",
        "A6": "TIN :",
        "A7": "OWNER'S NAME :",
        "A8": "OWNER'S TRADE NAME :",
        "A9": "OWNER'S ADDRESS :",
        "A11": "TAXABLE",
        "A12": "MONTH",
        "A14": "(1)",
        "B11": "TAXPAYER",
        "B12": "IDENTIFICATION",
        "B13": "NUMBER",
        "B14": "(2)",
        "C11": "REGISTERED NAME",
        "C14": "(3)",
        "D11": "NAME OF SUPPLIER",
        "D12": "(Last Name, First Name, Middle Name)",
        "D14": "(4)",
        "E11": "SUPPLIER'S ADDRESS",
        "E14": "(5)",
        "F11": "AMOUNT OF",
        "F12": "GROSS PURCHASE",
        "F14": "(6)",
        "G11": "AMOUNT OF",
        "G12": "EXEMPT PURCHASE",
        "G14": "(7)",
        "H11": "AMOUNT OF",
        "H12": "ZERO-RATED PURCHASE",
        "H14": "(8)",
        "I11": "AMOUNT OF",
        "I12": "TAXABLE PURCHASE",
        "I14": "(9)",
        "J11": "AMOUNT OF",
        "J12": "PURCHASE OF SERVICES",
        "J14": "(10)",
        "K11": "AMOUNT OF",
        "K12": "PURCHASE OF CAPITAL GOODS",
        "K14": "(11)",
        "L11": "AMOUNT OF",
        "L12": "PURCHASE OF GOODS OTHER THAN CAPITAL GOODS",
        "L14": "(12)",
        "M11": "AMOUNT OF",
        "M12": "INPUT TAX",
        "M14": "(13)",
        "N11": "AMOUNT OF",
        "N12": "GROSS TAXABLE PURCHASE",
        "N14": "(14)"
    }

    for cell, value in header_cells.items():
        ws[cell] = value
        ws[cell].font = Font(bold=True, size=10)

    # Column widths
    col_widths = {
        "A": 25, "B": 25, "C": 50, "D": 50, "E": 50,
        "F": 25, "G": 25, "H": 25, "I": 25,
        "J": 25, "K": 25, "L": 25, "M": 25, "N": 25
    }
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    # Create accounting style
    accounting_style = NamedStyle(name="accounting_style",
                                  number_format='#,##0.00_);[Red](#,##0.00)')
    try:
        wb.add_named_style(accounting_style)
    except Exception:
        pass

    normal_font = Font(size=10)
    right_align = Alignment(horizontal='right')
    bold_font = Font(bold=True, size=10)

    # Write data rows
    start_row = 15
    for r_idx, row in enumerate(processed, start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            if c_idx == 1:  # Column A (taxable month)
                cell.alignment = right_align
            if c_idx >= 6:  # Columns F through N (numeric columns)
                cell.number_format = '#,##0.00'

    last_data_row = start_row + len(processed) - 1
    
    # Add Grand Total row
    grand_total_row = last_data_row + 2
    
    # Write "Grand Total:" label in column A
    ws.cell(row=grand_total_row, column=1, value="Grand Total :")
    ws.cell(row=grand_total_row, column=1).font = bold_font
    ws.cell(row=grand_total_row, column=1).alignment = Alignment(horizontal='right')
    
    # Write totals for columns F through N
    column_letters = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
    for i, col_letter in enumerate(column_letters):
        cell = ws.cell(row=grand_total_row, column=6 + i, value=column_totals[i])
        cell.number_format = '#,##0.00'
        cell.font = bold_font
    
    # Add "END OF REPORT"
    end_of_report_row = grand_total_row + 2
    ws.cell(row=end_of_report_row, column=1, value="END OF REPORT")
    ws.cell(row=end_of_report_row, column=1).font = Font(size=10)

    # Save to bytes
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name="Excel Report.xlsx"
    )

@main.route('/convert_sales_xlsx', methods=['POST'])
def convert_sales_xlsx():
    payload = request.get_json() or {}
    rows = payload.get('rows') or []
    if not rows:
        return jsonify({'error': 'No rows provided'}), 400

    processed = []
    for idx, line in enumerate(rows):
        if idx == 0:
            continue
        cols = [c.strip().strip('"') for c in line.split(',')]
        if len(cols) < 14:
            continue
        try:
            numeric_values = [(float(cols[i]) if cols[i] else 0.0) for i in range(9, 14)]
        except Exception:
            numeric_values = [0.0] * 6

        formatted_row = [
            cols[14] if len(cols) > 14 else '',
            format_taxpayer_id(cols[2]) if len(cols) > 2 else '',
            cols[3] if len(cols) > 3 else '',
            f"{cols[4]}, {cols[5]} {cols[6]}",
            f"{cols[7]} {cols[8]}",
            sum(numeric_values[0:3]),
            numeric_values[0],
            numeric_values[1],
            numeric_values[2],
            numeric_values[3],
            sum(numeric_values[2:4])
        ]
        processed.append(formatted_row)

    if not processed:
        return jsonify({'error': 'No valid data to convert'}), 400

    wb = openpyxl.Workbook()
    ws = wb.active

    header_cells = {
        "A1": "SALES TRANSACTION",
        "A2": "RECONCILIATION OF LISTING FOR ENFORCEMENT",
        "A6": "TIN :",
        "A7": "OWNER'S NAME :",
        "A8": "OWNER'S TRADE NAME :",
        "A9": "OWNER'S ADDRESS :",
        "A11": "TAXABLE",
        "A12": "MONTH",
        "A14": "(1)",
        "B11": "TAXPAYER",
        "B12": "IDENTIFICATION",
        "B13": "NUMBER",
        "B14": "(2)",
        "C11": "REGISTERED NAME",
        "C14": "(3)",
        "D11": "NAME OF CUSTOMER",
        "D12": "(Last Name, First Name, Middle Name)",
        "D14": "(4)",
        "E11": "CUSTOMER'S ADDRESS",
        "E14": "(5)",
        "F11": "AMOUNT OF",
        "F12": "GROSS SALES",
        "F14": "(6)",
        "G11": "AMOUNT OF",
        "G12": "EXEMPT SALES",
        "G14": "(7)",
        "H11": "AMOUNT OF",
        "H12": "ZERO RATED SALES",
        "H14": "(8)",
        "I11": "AMOUNT OF",
        "I12": "TAXABLE SALES",
        "I14": "(9)",
        "J11": "AMOUNT OF",
        "J12": "OUTPUT TAX",
        "J14": "(10)",
        "K11": "AMOUNT OF",
        "K12": "GROSS TAXABLE SALES",
        "K14": "(11)",
    }

    for cell, value in header_cells.items():
        ws[cell] = value
        ws[cell].font = Font(bold=True, size=10)

    col_widths = {"A": 25, "B": 25, "C": 50, "D": 50, "E": 50,
                  "F": 25, "G": 25, "H": 25, "I": 25,
                  "J": 25, "K": 25}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w

    accounting_style = NamedStyle(name="accounting_style",
                                  number_format='#,##0.00_);[Red](#,##0.00)')
    try:
        wb.add_named_style(accounting_style)
    except Exception:
        pass

    normal_font = Font(size=10)
    right_align = Alignment(horizontal='right')

    start_row = 15
    for r_idx, row in enumerate(processed, start=start_row):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = normal_font
            if c_idx == 1:
                cell.alignment = right_align
            if c_idx >= 6:
                cell.number_format = '#,##0.00'

    last_data_row = start_row + len(processed) - 1
    grand_total_row = last_data_row + 2
    end_of_report_row = grand_total_row + 2

    ws[f"A{grand_total_row}"] = "Grand Total :"
    ws[f"A{grand_total_row}"].font = Font(size=10)

    for col in list('FGHIJK'):
        ws[f"{col}{grand_total_row}"] = f"=SUM({col}{start_row}:{col}{last_data_row})"
        ws[f"{col}{grand_total_row}"].number_format = '#,##0.00'
        ws[f"{col}{grand_total_row}"].font = Font(bold=True, size=10)

    ws[f"A{end_of_report_row}"] = "END OF REPORT"
    ws[f"A{end_of_report_row}"].font = Font(size=10)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name="Excel Report.xlsx"
    )